'''
TODO:
- make item info nicer. New window with empty searchbar at the top
  starts out with a scrollbar listing all items. As user types, search gets refined
  ex: typing octavia returns all items in the octavia set

- enable/disable wanting an item with a toggleable button on the wanted status

- add to wantlist

- view list

- implement live price checking

- scrollbar on main page

- owned increase/decrease buttons

- search for bugs

- error messages

- general styling

- color coding values

- icons for ducats/plat and maybe buttons at the top (play button for run, etc.)

- BUGLIST:

    * pressing cancel/x on the search item box
    * unable to read screen

'''




import pyscreenshot
import PIL.Image
import win32gui
import time
import pytesseract
from difflib import SequenceMatcher
from collections import namedtuple
import xlrd
import openpyxl
import os
from colorama import Fore, Style
from tkinter import *
import tkinter.messagebox
import tkinter.simpledialog
import re

# similar to C structs for holding item info. Maybe a class would be better
itemStruct = namedtuple("itemStruct", "name need ducat plat vaulted x y owned")

# holds a list of itemstructs
dropTable = []


WANT_COLOR = '#FFDC00'
EXIT_COLOR = '#FF4136'
UNWANT_COLOR = '#DDDDDD'


# gets info from the spreadsheet and stores it in the dropTable list
def readSheet():
    dropTable.clear()
    wb = openpyxl.load_workbook('warSpread.xlsx')
    ws = wb.active

    # iterate through all the items in the spreadsheet column and make items out of them
    for i, row in enumerate(ws.iter_rows()):
        tempItem = itemStruct(name=str(ws.cell(i + 2, 1).value).upper(),
                              need=str(ws.cell(i + 2, 3).value),
                              ducat=str(ws.cell(i + 2, 4).value),
                              plat=str(ws.cell(i + 2, 5).value),
                              vaulted=str(ws.cell(i + 2, 6).value),
                              owned=str(ws.cell(i + 2, 7).value),
                              x=str(i + 2),
                              y=str(3))

        # add the item to the array
        dropTable.append(tempItem)

    # deleting the last item since it is just an extra value in the spreadsheet for readability
    del dropTable[(len(dropTable) - 1)]


# takes a string and return the item whose name most closely matches the string, returns "nf" if not found
def searchByName(itemName):
    readSheet()
    if(itemName == "ALL"):
        return dropTable
    max = 0
    closestItem = None
    for i, item in enumerate(dropTable):
        if item.name == itemName:
            return item
    for i, item in enumerate(dropTable):
        if similar(itemName, item.name) > max:
            max = similar(itemName, item.name)
            closestItem = item
    if closestItem == None:
        return "nf"
    else:
        return closestItem


# gets information on a given item name string
def infoCommand(searchName):
    readSheet()
    if searchByName(searchName) == "nf":
        return
    tempItem = searchByName(searchName)
    print("Received information for item: " + tempItem.name + "...\n")
    print("Ducat value: " + tempItem.ducat)
    print("Plat value: " + tempItem.plat)
    if tempItem.vaulted == "V":
        print("Is item vaulted: YES")
    else:
        print("Is item vaulted: NO")
    print("Item on wantlist: " + tempItem.need)


# takes a string and adds that item to the wantlist
def wantCommand(searchName):
    readSheet()
    if searchByName(searchName) == "nf":
        return
    tempItem = searchByName(searchName)
    print(tempItem)
    if tempItem.need == "Y":
        wb = openpyxl.load_workbook('warSpread.xlsx')
        # wb.save('document.xlsx', as_template=False)
        ws = wb.active
        ws.cell(int(tempItem.x), int(tempItem.y), "N")
        wb.save("warSpread.xlsx")
    elif tempItem.need == "N":
        wb = openpyxl.load_workbook('warSpread.xlsx')
        # wb.save('document.xlsx', as_template=False)
        ws = wb.active
        ws.cell(int(tempItem.x), int(tempItem.y), "Y")
        wb.save("warSpread.xlsx")
    readSheet()
    tempItem = searchByName(searchName)
    print(tempItem)



# displays all items in the wantlist
def wantlistCommand():
    for i, item in enumerate(dropTable):
        if item.need == "Y":
            print(item.name)


# helper function used to find the most similar string from a list
def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()


def readScreen():
    readSheet()
    # grab the part of the screen we are interested in.
    # These cordinates are for my monitor which has a resolution of
    # 2560x1080, you will need to change these to match your monitor
    screenGrab = pyscreenshot.grab(bbox=(425, 454, 2135, 490))
    # save the screengrab as a png so we can open it with pytesseract
    screenGrab.save('screenshot2.png')
    fp = open("screenshot2.png", "rb")
    # we are using PIL since that is the type that tesseract expects
    img = PIL.Image.open(fp)
    # screen text is the string version of the picture
    # it is very common for it to pick up artifacts, so we need to clean it.
    screenText = pytesseract.image_to_string(img)

    # print statement to show us what was read in console.
    print("\nScreen text: " + screenText+"\n")

    # using regular expressions to clean up any artifacts
    # no item contains any non alphabetic characters, so remove any that were picked up
    regex = re.compile('[^a-zA-Z ]')
    screenText = regex.sub('', screenText)
    screenText = re.sub(' +', ' ', screenText)

    # print the cleaned text so we can look at it
    print("\nCleaned Screen text: " + screenText+"\n")

    # split the text by spaces into an array
    screenText = screenText.split(" ")

    # printing the split text for reading
    print("\nSplit Screen text: ")
    print(screenText)
    print("\n")

    # go through all the words in the array and remove them if they have less than three letters
    # no item has a word three letters or less, so if they exist, they are an error with the OCR
    for i, items in enumerate(screenText):
        if len(items) <= 3:
            screenText.pop(i)

    # print our final array of words
    print("\n final arrauy:\n")
    print(screenText)
    print("\n")

    # we remove all instances of "FORMA" since we are not interested in those items
    itemsOnScreen = []
    for i, word in enumerate(screenText):
        # we use similar since it is common for the screen capture to get the spelling wrong
        if similar(word.upper(), "FORMA") > 0.75:
            screenText.pop(i)

    # every item has the word prime as its second word. We can use thie knowledge to find out
    # where each item name starts and ends since not all items have the same number of words
    for i, word in enumerate(screenText):
        if similar(word.upper(), "PRIME") > 0.75:
            tempItemString = ""
            tempItemString += screenText[i-1]+" "
            tempItemString += screenText[i]+" "
            tempItemString += screenText[i+1]
            # tempstring now contains enough of the item name for us to recognize it in a search
            itemsOnScreen.append(tempItemString)

    print(itemsOnScreen)

    # now we go through the strings and see if we can find matching items for them
    tempItems = []
    for i, item in enumerate(itemsOnScreen):
        tempItem = searchByName(item)
        tempItems.append(tempItem)

    # now, we should have an array of items that are the closest match to the screen text
    print("Temp Items: ")
    print(tempItems)

    return tempItems
    # print(text)

def getItemsContaining(searchKey):
    readSheet()
    matching = []
    for items in dropTable:
        if items.name.contains(searchKey):
            matching.append(items)
    return matching

def toggleWant(buttn, item):
    wantCommand(item.name)
    newItem = searchByName(item.name)
    if newItem.need == "Y":
        buttn.config(bg=WANT_COLOR, text="Wanted: Y")
    else:
        buttn.config(bg=UNWANT_COLOR, text="Wanted: N")


def addItem(screen, item):
    readSheet()
    itemFrame = Frame(screen)

    itemFrame.pack(anchor="nw", fill=X)

    separatorFrame = Frame(itemFrame, bg="black", pady=3)
    separatorFrame.pack(anchor="n", fill=X)

    itemNameFrame = Frame(itemFrame)
    itemNameFrame.pack(anchor="nw")
    itemNameLabel = Label(itemNameFrame, text=item.name)
    itemNameLabel.config(font='Calibri 17 bold', padx=15, pady=5)
    itemNameLabel.pack(anchor="w")

    infoFrame = Frame(itemFrame)
    infoFrame.pack(anchor="nw", padx=15)

    ducatLabel = Label(infoFrame, text="Ducats: " + item.ducat)
    ducatLabel.config(font='Calibri 12 bold', padx=15, pady=5)
    ducatLabel.pack(side=LEFT)

    platLabel = Label(infoFrame, text="Platinum: " + item.plat)
    platLabel.config(font='Calibri 12 bold', padx=15, pady=5)
    platLabel.pack(side=LEFT)

    wantButton = Button(infoFrame, text="Wanted: " + item.need)
    wantButton.config(font='Calibri 12 bold', padx=15, pady=5, command=lambda arg=wantButton:toggleWant(arg, item))
    if item.need == "Y":
        wantButton.config(bg=WANT_COLOR)
    else:
        wantButton.config(bg=UNWANT_COLOR)
    wantButton.pack(side=LEFT)

    vaultLabel = Label(infoFrame, text="Vaulted: " + item.vaulted)
    vaultLabel.config(font='Calibri 12 bold', padx=15, pady=5)
    vaultLabel.pack(side=LEFT)

    ownedFrame = Frame(itemFrame)
    ownedFrame.pack(anchor="nw", padx=15)

    ownedLabel = Label(ownedFrame, text="Number Owned: ")
    ownedLabel.config(font='Calibri 12 bold', padx=15, pady=5)
    ownedLabel.pack(side=LEFT)

    decreaseButton = Button(ownedFrame, text="-")
    decreaseButton.config(font='Calibri 12 bold')
    decreaseButton.pack(side=LEFT, pady=5, padx=5)

    ownedLabel = Label(ownedFrame, text=" " + item.owned)
    ownedLabel.config(font='Calibri 12 bold', padx=15, pady=5)
    ownedLabel.pack(side=LEFT)

    increaseButton = Button(ownedFrame, text="+")
    increaseButton.config(font='Calibri 12 bold')
    increaseButton.pack(side=LEFT, pady=5, padx=5)

    removeButton = Button(infoFrame, text="X", command=lambda: itemFrame.destroy())
    removeButton.pack(anchor="ne", pady=5, padx=5)
    removeButton.config(font='Calibri 12 bold', bg="red")


class MainWindow(Frame):
    counter = 0

    def __init__(self, *args, **kwargs):
        Frame.__init__(self, *args, **kwargs)

        '''
        self.listButton = tk.Button(self, text="View wantlist", command=self.create_window)
        self.listButton.pack(side="top")
        '''

        self.toolBarFrame = Frame(root)
        self.toolBarFrame.pack(anchor="nw", fill=X)

        self.runButton = Button(self.toolBarFrame, text="RUN", command=self.runScreenGrab)
        self.runButton.config(font='Calibri 12 bold')
        self.runButton.pack(side=RIGHT, pady=10, padx=20)

        self.addButton = Button(self.toolBarFrame, text="Add to Wantlist")
        self.addButton.config(font='Calibri 12 bold')
        self.addButton.pack(side=LEFT, pady=10, padx=2)

        self.searchButton = Button(self.toolBarFrame, text="Search Item", command=self.infoWindow)
        self.searchButton.config(font='Calibri 12 bold')
        self.searchButton.pack(side=LEFT, pady=10, padx=2)

        self.wantListButton = Button(self.toolBarFrame, text="View List", command=self.listWindow)
        self.wantListButton.config(font='Calibri 12 bold')
        self.wantListButton.pack(side=LEFT, pady=10, padx=2)



    def create_window(self):
        self.counter += 1
        t = Toplevel(self)
        t.wm_title("Window #%s" % self.counter)
        l = Label(t, text="This is window #%s" % self.counter)
        l.pack(side="top", fill="both", expand=True, padx=100, pady=100)

    def listWindow(self):
        readSheet()
        t = Toplevel(self)
        t.wm_title("Item Info")
        t.minsize(width=500, height=200)
        for item in dropTable:
            if item.need == "Y":
                addItem(t, item)

    def infoWindow(self):
        itemName = tkinter.simpledialog.askstring("Name prompt", "enter your name")
        itemName = itemName.upper()
        t = Toplevel(self)
        t.wm_title("Item Info")
        t.minsize(width=500, height=200)
        item = searchByName(itemName)
        addItem(t, item)
        '''
        t = Toplevel(self)
        t.wm_title("Item Info")
        t.minsize(width=500, height=200)
        searchFrame = Frame(t)
        searchFrame.pack(fill=X, padx=10, pady=5)

        searchBox = Entry(searchFrame)
        searchBox.insert(0, "Enter Item Name")
        searchBox.config(font="Calibri 14 bold")
        searchBox.pack(anchor="nw", fill=X, padx=15, pady=15)

        
        item = searchByName(itemName)
        addItem(t, item)
        '''

    def runScreenGrab(self):
        currentItems = readScreen()
        for i, item in enumerate(currentItems):
            addItem(self, item)


if __name__ == '__main__':
    readSheet()
    root = Tk()
    main = MainWindow(root)
    main.pack(side="top", fill="both", expand=True)
    root.mainloop()
    # var = tkinter.simpledialog.askstring("Name prompt", "enter your name")
